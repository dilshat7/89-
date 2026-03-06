import openpyxl

EXCEL_PATH = r"C:\Users\ddjab\OneDrive\Рабочий стол\Выгрузка товар\translations.xlsx"

DESC_RU     = "Размеры: S. M. L."
DESC_EN     = "size S. M. L."
DESC_UZ_CYR = "ўлчамлари: S. M. L."
DESC_UZ_LAT = "o'lchamlari: S. M. L."

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb["Описания"]

headers = [cell.value for cell in ws[1]]
print("Колонки:", headers)

col_ru  = headers.index("description (RU)") + 1
col_en  = headers.index("description_en (EN)") + 1
col_uz  = headers.index("description_uz (перевод)") + 1
col_cyr = headers.index("description_uz_cyr (перевод)") + 1

for row in range(2, ws.max_row + 1):
    ws.cell(row=row, column=col_ru,  value=DESC_RU)
    ws.cell(row=row, column=col_en,  value=DESC_EN)
    ws.cell(row=row, column=col_uz,  value=DESC_UZ_LAT)
    ws.cell(row=row, column=col_cyr, value=DESC_UZ_CYR)

wb.save(EXCEL_PATH)
print(f"Готово! {ws.max_row - 1} описаний обновлено.")
print(f"  RU:  {DESC_RU}")
print(f"  EN:  {DESC_EN}")
print(f"  UZ:  {DESC_UZ_LAT}")
print(f"  CYR: {DESC_UZ_CYR}")
