import re
import os

import openpyxl

EXCEL_PATH = r"C:\Users\ddjab\OneDrive\Рабочий стол\Выгрузка товар\translations.xlsx"

CATEGORY_MERGE = {
    ("Мужчины", "Верхняя одежда"):    "Мужская верхняя одежда",
    ("Мужчины", "Брюки"):             "Мужские брюки",
    ("Мужчины", "Шорты"):             "Мужские шорты",
    ("Мужчины", "Рубашки"):           "Мужские рубашки",
    ("Мужчины", "Футболки"):          "Мужские футболки",
    ("Мужчины", "Джинсы"):            "Мужские джинсы",
    ("Женщины", "Верхняя одежда"):    "Женская верхняя одежда",
    ("Женщины", "Брюки"):             "Женские брюки",
    ("Женщины", "Юбки"):              "Женские юбки",
    ("Женщины", "Блузки и рубашки"):  "Женские блузки и рубашки",
    ("Женщины", "Пиджаки и жакеты"): "Женские пиджаки и жакеты",
    ("Женщины", "Спортивные костюмы"):"Женские спортивные костюмы",
    ("Женщины", "Джинсы"):            "Женские джинсы",
    ("Женщины", "Шорты"):             "Женские шорты",
    ("Женщины", "Платья"):            "Женские платья",
    ("Женщины", "Свитера"):           "Женские свитера",
}

GENDER_NEUTRAL = {
    "ветровка": ("Мужская ветровка", "Женская ветровка"),
    "куртка":   ("Мужская куртка",   "Женская куртка"),
    "пальто":   ("Мужское пальто",   "Женское пальто"),
    "пуховик":  ("Мужской пуховик",  "Женский пуховик"),
    "тренч":    ("Мужской тренч",    "Женский тренч"),
    "брюки":    ("Мужские брюки",    "Женские брюки"),
    "джинсы":   ("Мужские джинсы",   "Женские джинсы"),
    "шорты":    ("Мужские шорты",    "Женские шорты"),
    "рубашка":  ("Мужская рубашка",  "Женская рубашка"),
    "блузка":   ("Мужская блузка",   "Женская блузка"),
    "блузки":   ("Мужские блузки",   "Женские блузки"),
    "топ":      ("Мужской топ",      "Женский топ"),
    "пиджак":   ("Мужской пиджак",   "Женский пиджак"),
    "жакет":    ("Мужской жакет",    "Женский жакет"),
    "жилет":    ("Мужской жилет",    "Женский жилет"),
    "юбка":     ("Мужская юбка",     "Женская юбка"),
    "платье":   ("Мужское платье",   "Женское платье"),
    "свитер":   ("Мужской свитер",   "Женский свитер"),
    "кардиган": ("Мужской кардиган", "Женский кардиган"),
    "худи":     ("Мужское худи",     "Женское худи"),
    "свитшот":  ("Мужской свитшот",  "Женский свитшот"),
    "футболка": ("Мужская футболка", "Женская футболка"),
    "костюм":   ("Мужской костюм",   "Женский костюм"),
    "джоггеры": ("Мужские джоггеры", "Женские джоггеры"),
}

COLOR_EN = {
    "бежевый":                  "beige",
    "серый":                    "grey",
    "серый твид":               "grey tweed",
    "тёмно-серый":              "dark grey",
    "чёрный":                   "black",
    "белый":                    "white",
    "молочный":                 "cream",
    "коричневый":               "brown",
    "тёмно-коричневый":         "dark brown",
    "оливковый":                "olive",
    "хаки/оливковый":           "khaki/olive",
    "коричневый/оливковый":     "brown/olive",
    "тёмно-синий":              "navy",
    "голубой":                  "light blue",
    "синий":                    "blue",
    "сине-белый в полоску":     "blue white stripes",
    "голубой в полоску":        "light blue stripes",
    "тёмно-синий в полоску":    "navy stripes",
    "белый в полоску":          "white stripes",
    "розовый":                  "pink",
    "бордовый":                 "burgundy",
    "бордовый в клетку":        "burgundy plaid",
    "красный в клетку":         "red plaid",
    "серо-синий в клетку":      "grey blue plaid",
    "сине-оранжевый в клетку":  "blue orange plaid",
    "бежевый в полоску":        "beige stripes",
    "голубой с принтом":        "light blue print",
    "оранжевый":                "orange",
    "терракотовый":             "terracotta",
    "нет фото":                 "no photo",
}

EN_TYPES = {
    "ветровка": "windbreaker",
    "куртка":   "jacket",
    "пальто":   "coat",
    "пуховик":  "puffer jacket",
    "тренч":    "trench coat",
    "брюки":    "trousers",
    "джинсы":   "jeans",
    "шорты":    "shorts",
    "рубашка":  "shirt",
    "блузка":   "blouse",
    "блузки":   "blouses",
    "топ":      "top",
    "пиджак":   "blazer",
    "жакет":    "jacket",
    "жилет":    "vest",
    "юбка":     "skirt",
    "платье":   "dress",
    "свитер":   "sweater",
    "кардиган": "cardigan",
    "худи":     "hoodie",
    "свитшот":  "sweatshirt",
    "футболка": "t-shirt",
    "костюм":   "suit",
    "джоггеры": "joggers",
}

UZ_TYPES = {
    "ветровка": "ветровкаси",
    "куртка":   "курткаси",
    "пальто":   "пальтоси",
    "пуховик":  "пуховиги",
    "тренч":    "тренчи",
    "брюки":    "шими",
    "джинсы":   "жинсиси",
    "шорты":    "шорти",
    "рубашка":  "кўйлаги",
    "блузка":   "блузкаси",
    "блузки":   "блузкаси",
    "топ":      "топи",
    "пиджак":   "пиджаги",
    "жакет":    "жакети",
    "жилет":    "жилети",
    "юбка":     "юбкаси",
    "платье":   "кўйлаги",
    "свитер":   "свитери",
    "кардиган": "кардигани",
    "худи":     "худиси",
    "свитшот":  "свитшоти",
    "футболка": "футболкаси",
    "костюм":   "костюми",
    "джоггеры": "джоггерси",
}

COLOR_UZ = {
    "бежевый":                 "қумранг",
    "серый":                   "кулранг",
    "серый твид":              "кулранг твид",
    "тёмно-серый":             "тўқ кулранг",
    "чёрный":                  "қора",
    "белый":                   "оқ",
    "молочный":                "сутранг",
    "коричневый":              "жигарранг",
    "тёмно-коричневый":        "тўқ жигарранг",
    "оливковый":               "зайтунранг",
    "хаки/оливковый":          "хаки",
    "коричневый/оливковый":    "жигарранг/зайтунранг",
    "тёмно-синий":             "тўқ кўк",
    "голубой":                 "осмонранг",
    "синий":                   "кўк",
    "сине-белый в полоску":    "кўк-оқ йўл-йўл",
    "голубой в полоску":       "осмонранг йўл-йўл",
    "тёмно-синий в полоску":   "тўқ кўк йўл-йўл",
    "белый в полоску":         "оқ йўл-йўл",
    "розовый":                 "пушти",
    "бордовый":                "тўқ қизил",
    "бордовый в клетку":       "тўқ қизил катакли",
    "красный в клетку":        "қизил катакли",
    "серо-синий в клетку":     "кулранг-кўк катакли",
    "сине-оранжевый в клетку": "кўк-тўқсариқ катакли",
    "бежевый в полоску":       "қумранг йўл-йўл",
    "голубой с принтом":       "осмонранг принтли",
    "оранжевый":               "тўқсариқ",
    "терракотовый":            "терракот",
    "нет фото":                "",
}

def build_title(gender: str, title_ru: str, color: str) -> str:
    clean = re.sub(r'\s+[A-Z]{2}\d{2}[\w\-]+', '', title_ru).strip().lower()
    for kw, (male_form, female_form) in GENDER_NEUTRAL.items():
        if kw in clean:
            base = male_form if gender == "Мужчины" else female_form
            if color and color != "нет фото":
                return f"{base} {color}"
            return base
    words = clean.split()
    skip = {"мужские", "женские", "мужская", "женская", "мужской", "женской", "мужское", "женское"}
    clean_word = next((w for w in words if w not in skip), words[0] if words else title_ru)
    prefix = "Мужская" if gender == "Мужчины" else "Женская"
    if color and color != "нет фото":
        return f"{prefix} {clean_word} {color}"
    return f"{prefix} {clean_word}"

def build_title_en(gender: str, title_ru: str, color: str) -> str:
    clean = re.sub(r'\s+[A-Z]{2}\d{2}[\w\-]+', '', title_ru).strip().lower()
    prefix = "Men's" if gender == "Мужчины" else "Women's"
    color_en = COLOR_EN.get(color, color)
    for kw, en_type in EN_TYPES.items():
        if kw in clean:
            if color_en and color_en != "no photo":
                return f"{prefix} {en_type} {color_en}"
            return f"{prefix} {en_type}"
    clean_word = clean.split()[0] if clean else title_ru
    if color_en and color_en != "no photo":
        return f"{prefix} {clean_word} {color_en}"
    return f"{prefix} {clean_word}"

def build_title_uz_cyr(gender: str, title_ru: str, color: str) -> str:
    clean = re.sub(r'\s+[A-Z]{2}\d{2}[\w\-]+', '', title_ru).strip().lower()
    prefix = "Эркаклар" if gender == "Мужчины" else "Аёллар"
    color_uz = COLOR_UZ.get(color, color)
    for kw, uz_type in UZ_TYPES.items():
        if kw in clean:
            if color_uz:
                return f"{prefix} {uz_type} {color_uz}"
            return f"{prefix} {uz_type}"
    clean_word = clean.split()[0] if clean else title_ru
    if color_uz:
        return f"{prefix} {clean_word} {color_uz}"
    return f"{prefix} {clean_word}"

CYR_TO_LAT = [
    ("шч", "shch"), ("Шщ", "Shch"),
    ("नग", "ng"),  ("НГ", "Ng"),
    ("आ", "yo"),  ("आ", "Yo"),
    ("ж", "j"),   ("Ж", "J"),
    ("з", "z"),   ("З", "Z"),
    ("ш", "sh"),  ("Ш", "Sh"),
    ("ч", "ch"),  ("Ч", "Ch"),
    ("щ", "sh"),  ("Щ", "Sh"),
    ("ю", "yu"),  ("Ю", "Yu"),
    ("я", "ya"),  ("Я", "Ya"),
    ("ё", "yo"),  ("Ё", "Yo"),
    ("а", "a"),   ("А", "A"),
    ("б", "b"),   ("Б", "B"),
    ("в", "v"),   ("В", "V"),
    ("г", "g"),   ("Г", "G"),
    ("д", "d"),   ("Д", "D"),
    ("е", "e"),   ("Е", "E"),
    ("и", "i"),   ("И", "I"),
    ("й", "y"),   ("Й", "Y"),
    ("к", "k"),   ("К", "K"),
    ("л", "l"),   ("Л", "L"),
    ("м", "m"),   ("М", "M"),
    ("н", "n"),   ("Н", "N"),
    ("о", "o"),   ("О", "O"),
    ("п", "p"),   ("П", "P"),
    ("р", "r"),   ("Р", "R"),
    ("с", "s"),   ("С", "S"),
    ("т", "t"),   ("Т", "T"),
    ("у", "u"),   ("У", "U"),
    ("ф", "f"),   ("Ф", "F"),
    ("х", "x"),   ("Х", "X"),
    ("ц", "ts"),  ("Ц", "Ts"),
    ("ы", "i"),   ("Ы", "I"),
    ("э", "e"),   ("Э", "E"),
    ("ў", "o‘"),  ("Ў", "O‘"),
    ("қ", "q"),   ("Қ", "Q"),
    ("ғ", "g‘"),  ("Ғ", "G‘"),
    ("ҳ", "h"),   ("Ҳ", "H"),
    ("ъ", ""),    ("Ъ", ""),
    ("ь", ""),    ("Ь", ""),
]

def cyr_to_lat(text: str) -> str:
    for cyr, lat in CYR_TO_LAT:
        text = text.replace(cyr, lat)
    return text

def build_title_uz_lat(gender: str, title_ru: str, color: str) -> str:
    cyr = build_title_uz_cyr(gender, title_ru, color)
    return cyr_to_lat(cyr)

import json

PRODUCTS_JSON = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "products.json")

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb["Названия"]

headers = [c.value for c in ws[1]]
color_col = headers.index("Цвет") if "Цвет" in headers else None
num_col   = headers.index("№")    if "№"    in headers else 0

colors_by_num = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    num = row[num_col]
    if num and color_col is not None:
        colors_by_num[num] = row[color_col] or ""

with open(PRODUCTS_JSON, encoding="utf-8") as f:
    products_json = json.load(f)

ws.delete_rows(1, ws.max_row)
ws.append(["№", "Категория", "title (RU)", "title_en (EN)", "Цвет", "title_uz_cyr (перевод)", "title_uz (перевод)"])

print(f"{'№':<4} {'Категория':<30} {'RU':<35} {'UZ lat'}")
print("-" * 110)

for i, product in enumerate(products_json, 1):
    if i not in colors_by_num:
        continue

    category    = product.get("parent_category", "")
    subcategory = product.get("category", "")
    title_ru_orig = product.get("title", "")
    color = colors_by_num[i]

    merged_cat       = CATEGORY_MERGE.get((category, subcategory), f"{category} {subcategory}".strip())
    new_title        = build_title(category, title_ru_orig, color)
    new_title_en     = build_title_en(category, title_ru_orig, color)
    new_title_uz_cyr = build_title_uz_cyr(category, title_ru_orig, color)
    new_title_uz_lat = build_title_uz_lat(category, title_ru_orig, color)

    ws.append([i, merged_cat, new_title, new_title_en, color, new_title_uz_cyr, new_title_uz_lat])
    print(f"{i:<4} {merged_cat:<30} {new_title:<35} {new_title_uz_lat}")

wb.save(EXCEL_PATH)
print(f"\nГотово! Сохранено в {EXCEL_PATH}")
