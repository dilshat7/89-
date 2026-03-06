import json
import os
import time
import argparse
import requests
import openpyxl

BASE_URL   = "https://aichat.ikujo.com"
EMAIL      = "sale.fix@hotmail.com"
PASSWORD   = "Password1!"
ACCOUNT_ID = 71

PRODUCTS_JSON = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "products.json")
EXCEL_PATH    = r"C:\Users\ddjab\OneDrive\Рабочий стол\Выгрузка товар\translations.xlsx"

CATEGORY_NAMES = {
    "Мужчины": {"en": "Men",   "ru": "Мужчины", "uz": "Erkaklar", "uz_cyr": "Эркаклар"},
    "Женщины": {"en": "Women", "ru": "Женщины", "uz": "Ayollar",  "uz_cyr": "Аёллар"},
}

SUBCATEGORY_NAMES = {
    ("Мужчины", "Верхняя одежда"):    {"en": "Men's Outerwear",        "ru": "Мужская верхняя одежда",    "uz": "Erkaklar tashqi kiyimlari",          "uz_cyr": "Эркаклар ташқи кийимлари"},
    ("Мужчины", "Брюки"):             {"en": "Men's Trousers",          "ru": "Мужские брюки",             "uz": "Erkaklar shimlari",                  "uz_cyr": "Эркаклар шимлари"},
    ("Мужчины", "Шорты"):             {"en": "Men's Shorts",            "ru": "Мужские шорты",             "uz": "Erkaklar shortlari",                 "uz_cyr": "Эркаклар шортлари"},
    ("Мужчины", "Рубашки"):           {"en": "Men's Shirts",            "ru": "Мужские рубашки",           "uz": "Erkaklar ko'ylaklari",               "uz_cyr": "Эркаклар кўйлаклари"},
    ("Мужчины", "Джинсы"):            {"en": "Men's Jeans",             "ru": "Мужские джинсы",            "uz": "Erkaklar jinsilari",                 "uz_cyr": "Эркаклар жинсилари"},
    ("Женщины", "Верхняя одежда"):    {"en": "Women's Outerwear",       "ru": "Женская верхняя одежда",    "uz": "Ayollar tashqi kiyimlari",           "uz_cyr": "Аёллар ташқи кийимлари"},
    ("Женщины", "Пиджаки и жакеты"): {"en": "Women's Blazers & Jackets","ru": "Женские пиджаки и жакеты", "uz": "Ayollar blazerlari va jaketlari",     "uz_cyr": "Аёллар блейзерлари ва жакетлари"},
    ("Женщины", "Юбки"):              {"en": "Women's Skirts",          "ru": "Женские юбки",              "uz": "Ayollar yubkalari",                  "uz_cyr": "Аёллар юбкалари"},
    ("Женщины", "Блузки и рубашки"):  {"en": "Women's Blouses & Shirts","ru": "Женские блузки и рубашки",  "uz": "Ayollar bluzalari va ko'ylaklari",   "uz_cyr": "Аёллар блузалари ва кўйлаклари"},
    ("Женщины", "Брюки"):             {"en": "Women's Trousers",        "ru": "Женские брюки",             "uz": "Ayollar shimlari",                   "uz_cyr": "Аёллар шимлари"},
    ("Женщины", "Шорты"):             {"en": "Women's Shorts",          "ru": "Женские шорты",             "uz": "Ayollar shortlari",                  "uz_cyr": "Аёллар шортлари"},
    ("Женщины", "Джинсы"):            {"en": "Women's Jeans",           "ru": "Женские джинсы",            "uz": "Ayollar jinsilari",                  "uz_cyr": "Аёллар жинсилари"},
    ("Женщины", "Спортивные костюмы"):{"en": "Women's Sportswear",      "ru": "Женские спортивные костюмы","uz": "Ayollar sport kostyumlari",          "uz_cyr": "Аёллар спорт костюмлари"},
    ("Женщины", "Платья"):            {"en": "Women's Dresses",         "ru": "Женские платья",            "uz": "Ayollar ko'ylaklari",                "uz_cyr": "Аёллар кўйлаклари"},
    ("Женщины", "Свитера"):           {"en": "Women's Sweaters",        "ru": "Женские свитера",           "uz": "Ayollar sviterlari",                 "uz_cyr": "Аёллар свитерлари"},
}


def load_excel_data():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # --- Названия ---
    ws1 = wb["Названия"]
    h1 = [c.value for c in ws1[1]]
    titles = {}
    for row in ws1.iter_rows(min_row=2, values_only=True):
        num = row[0]
        if not num:
            continue
        titles[num] = {
            "title_ru":      row[h1.index("title (RU)")],
            "title_en":      row[h1.index("title_en (EN)")],
            "title_uz_cyr":  row[h1.index("title_uz_cyr (перевод)")] if "title_uz_cyr (перевод)" in h1 else None,
            "title_uz":      row[h1.index("title_uz (перевод)")]     if "title_uz (перевод)"     in h1 else None,
        }

    # --- Описания ---
    ws2 = wb["Описания"]
    h2 = [c.value for c in ws2[1]]
    descs = {}
    for row in ws2.iter_rows(min_row=2, values_only=True):
        num = row[0]
        if not num:
            continue
        descs[num] = {
            "desc_ru":      row[h2.index("description (RU)")],
            "desc_en":      row[h2.index("description_en (EN)")],
            "desc_uz":      row[h2.index("description_uz (перевод)")],
            "desc_uz_cyr":  row[h2.index("description_uz_cyr (перевод)")],
            "parent_cat":   row[h2.index("Категория")],
            "child_cat":    row[h2.index("Подкатегория")],
        }

    return titles, descs


class IkujoUploader:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})

    def login(self):
        print("Авторизация...")
        url = f"{BASE_URL}/auth/sign_in"
        resp = self.session.post(url, json={"email": EMAIL, "password": PASSWORD})
        resp.raise_for_status()
        data = resp.json()

        inner = data.get("data", data)
        token = inner.get("access_token") or inner.get("token")

        if "access-token" in resp.headers:
            self.session.headers.update({
                "access-token": resp.headers["access-token"],
                "client":       resp.headers["client"],
                "uid":          resp.headers["uid"],
            })
            del self.session.headers["Content-Type"]
            print("✅ Авторизован (devise token)")
        elif token:
            self.session.headers.update({"Authorization": f"Bearer {token}"})
            del self.session.headers["Content-Type"]
            print("✅ Авторизован (Bearer)")
        else:
            print(f"❌ Токен не получен: {data}")
            return False
        return True

    def get_categories(self):
        url = f"{BASE_URL}/api/v1/accounts/{ACCOUNT_ID}/product_categories"
        resp = self.session.get(url, params={"per_page": 200})
        resp.raise_for_status()
        data = resp.json()
        return data if isinstance(data, list) else data.get("data", [])

    def create_category(self, name_ru, parent_id=None, parent_name=None):
        if parent_id and parent_name:
            names = SUBCATEGORY_NAMES.get((parent_name, name_ru), {
                "en": name_ru, "ru": name_ru, "uz": name_ru, "uz_cyr": name_ru
            })
        else:
            names = CATEGORY_NAMES.get(name_ru, {
                "en": name_ru, "ru": name_ru, "uz": name_ru, "uz_cyr": name_ru
            })
        url = f"{BASE_URL}/api/v1/accounts/{ACCOUNT_ID}/product_categories"
        payload = {
            "name": {
                "en":     names["en"],
                "ru":     names["ru"],
                "uz":     names["uz"],
                "uz_cyr": names["uz_cyr"],
            }
        }
        if parent_id:
            payload["parent_id"] = parent_id
        resp = self.session.post(url, json=payload)
        resp.raise_for_status()
        cat = resp.json()
        print(f"  ✅ Категория: '{names['en']}' (ID={cat.get('id')})")
        return cat

    def build_category_map(self):
        print("\nПолучаем категории...")
        existing = self.get_categories()

        existing_ru = {}
        for cat in existing:
            name_obj = cat.get("name", {})
            name_ru = (name_obj.get("ru") or name_obj.get("en") or "").strip() if isinstance(name_obj, dict) else str(name_obj).strip()
            existing_ru[name_ru] = cat["id"]

        category_map = {}

        with open(PRODUCTS_JSON, encoding="utf-8") as f:
            products = json.load(f)

        pairs = set()
        for p in products:
            parent = p.get("parent_category", "")
            child  = p.get("category", "")
            if parent and child:
                pairs.add((parent, child))

        for parent_name, child_name in sorted(pairs):
            names = SUBCATEGORY_NAMES.get((parent_name, child_name), {
                "en": f"{parent_name} {child_name}",
                "ru": f"{parent_name} {child_name}",
                "uz": f"{parent_name} {child_name}",
                "uz_cyr": f"{parent_name} {child_name}",
            })
            expected_ru = names["ru"]

            if expected_ru in existing_ru:
                found_id = existing_ru[expected_ru]
                print(f"  ℹ️  '{expected_ru}' уже есть (ID={found_id})")
                category_map[(parent_name, child_name)] = found_id
            else:
                url = f"{BASE_URL}/api/v1/accounts/{ACCOUNT_ID}/product_categories"
                payload = {"name": {"en": names["en"], "ru": names["ru"], "uz": names["uz"], "uz_cyr": names["uz_cyr"]}}
                resp = self.session.post(url, json=payload)
                resp.raise_for_status()
                cat = resp.json()
                cat_id = cat.get("id")
                print(f"  ✅ Категория: '{names['en']}' (ID={cat_id})")
                category_map[(parent_name, child_name)] = cat_id

        return category_map

    def _extract_price(self, price_str):
        import re
        if isinstance(price_str, (int, float)):
            return int(price_str)
        numbers = re.findall(r"\d+", str(price_str).replace(" ", ""))
        return int("".join(numbers)) if numbers else 0

    def create_product(self, product, t, d, category_id, index, total):
        title_en  = t.get("title_en")  or product.get("title_en", product.get("title", ""))
        title_ru  = t.get("title_ru")  or product.get("title", "")
        title_uz  = t.get("title_uz")  or title_ru
        title_cyr = t.get("title_uz_cyr") or title_ru

        desc_en  = d.get("desc_en")     or "size S. M. L."
        desc_ru  = d.get("desc_ru")     or "Размеры: S. M. L."
        desc_uz  = d.get("desc_uz")     or "o'lchamlari: S. M. L."
        desc_cyr = d.get("desc_uz_cyr") or "ўлчамлари: S. M. L."

        print(f"[{index}/{total}] {title_en[:45]}", end=" ")

        url = f"{BASE_URL}/api/v1/accounts/{ACCOUNT_ID}/products"
        form_data = {
            # Порядок: EN → RU → UZ латиница → UZ кириллица
            "product[name][en]":          title_en,
            "product[name][ru]":          title_ru,
            "product[name][uz]":          title_uz,
            "product[name][uz_cyr]":      title_cyr,
            "product[description][en]":   desc_en,
            "product[description][ru]":   desc_ru,
            "product[description][uz]":   desc_uz,
            "product[description][uz_cyr]": desc_cyr,
            "product[product_category_id]": str(category_id),
            "product[quantity]":          str(product.get("quantity", 1)),
            "product[track_quantity]":    "true",
            "product[is_published]":      "true",
            "product[prices_attributes][0][price]":         str(self._extract_price(product.get("price", 0))),
            "product[prices_attributes][0][currency_code]": "UZS",
        }

        local_images = product.get("local_images", [])
        if not local_images and product.get("local_image"):
            local_images = [product["local_image"]]

        local_images = local_images[:5]

        files = []
        buffers = []
        for img_path in local_images:
            if not os.path.exists(img_path):
                continue
            ext = os.path.splitext(img_path)[1].lower()
            fname = os.path.basename(img_path)
            if ext == ".webp":
                from PIL import Image
                import io
                img = Image.open(img_path).convert("RGB")
                buf = io.BytesIO()
                img.save(buf, format="JPEG", quality=90)
                buf.seek(0)
                buffers.append(buf)
                fname = fname.replace(".webp", ".jpg")
                files.append(("product[images][]", (fname, buf, "image/jpeg")))
            else:
                f = open(img_path, "rb")
                buffers.append(f)
                mime = {"image/jpeg": "image/jpeg", ".jpg": "image/jpeg",
                        ".jpeg": "image/jpeg", ".png": "image/png"}.get(ext, "image/jpeg")
                files.append(("product[images][]", (fname, f, "image/jpeg")))

        print(f"({len(files)} фото)", end=" ")

        try:
            resp = self.session.post(url, data=form_data, files=files)
            if resp.status_code not in (200, 201):
                print(f"❌ HTTP {resp.status_code}: {resp.text[:200]}")
                return None
            prod = resp.json()
            print(f"✅ ID={prod.get('id')}")
            return prod
        except Exception as e:
            print(f"❌ {e}")
            return None
        finally:
            for f in buffers:
                f.close()

    def run(self, limit=None):
        if not self.login():
            return

        category_map = self.build_category_map()
        print(f"\n✅ Категорий создано/найдено: {len(category_map)}")

        titles, descs = load_excel_data()
        print(f"📊 Загружено из Excel: {len(titles)} товаров")

        with open(PRODUCTS_JSON, encoding="utf-8") as f:
            products = json.load(f)

        SKIP_CATEGORIES = {
            ("Мужчины", "Аксессуары"),
        }

        products_to_upload = [
            (i + 1, p) for i, p in enumerate(products)
            if (i + 1) in titles and (i + 1) in descs
            and (p.get("parent_category", ""), p.get("category", "")) not in SKIP_CATEGORIES
        ]

        if limit:
            products_to_upload = products_to_upload[:limit]

        total = len(products_to_upload)
        print(f"\n{'='*60}")
        print(f"Загружаем {total} товаров...")
        print(f"{'='*60}\n")

        ok = fail = 0

        for idx, (num, product) in enumerate(products_to_upload, 1):
            t = titles[num]
            d = descs[num]

            parent_name = d.get("parent_cat") or product.get("parent_category", "")
            child_name  = d.get("child_cat")  or product.get("category", "")
            key = (parent_name, child_name)

            category_id = category_map.get(key)
            if not category_id:
                print(f"[{idx}/{total}] ❌ Категория не найдена: {key}")
                fail += 1
                continue

            result = self.create_product(product, t, d, category_id, idx, total)
            if result:
                ok += 1
            else:
                fail += 1

            time.sleep(0.5)

        print(f"\n{'='*60}")
        print(f"✅ Успешно: {ok}")
        print(f"❌ Ошибки:  {fail}")
        print(f"{'='*60}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--test", type=int, default=None, help="Загрузить только N товаров для теста")
    args = parser.parse_args()

    uploader = IkujoUploader()
    uploader.run(limit=args.test)
